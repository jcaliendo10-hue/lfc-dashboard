#!/usr/bin/env python3
"""
Liverpool FC dashboard — data updater.

Scrapes Liverpool news from across the web (direct outlet RSS + Google News
search feeds that aggregate hundreds of sources and journalists, including
dedicated queries for trusted reporters who break news on social media —
Fabrizio Romano, James Pearce, Paul Joyce, David Ornstein, etc.), scores each
story by (1) SOURCE CREDIBILITY/ACCURACY and (2) RELEVANCE, ranks them, and writes:
  - data/news.json           (consumed by the dashboard)
  - data/liverpool_data.xlsx (the Excel "backend" / source of truth)

Run locally:  python scripts/update_data.py
Runs daily on GitHub Actions (see .github/workflows/update.yml).

No API keys required — everything is public RSS.
"""

import json
import re
import sys
import datetime as dt
from pathlib import Path
from html import unescape
from urllib.parse import quote_plus

try:
    import feedparser
except ImportError:
    sys.exit("Missing dependency: pip install feedparser openpyxl")

ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
DATA.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# 1. SOURCE CREDIBILITY / ACCURACY MAP
# A reliability score (1-10) per outlet/domain. Used to weight every story,
# whether it comes from a direct feed or is surfaced via Google News.
# Tier 1 (9-10): primary, rarely-wrong sources & top reporters' outlets.
# Tier 2 (7-8):  strong national media.
# Tier 3 (5-6):  reputable fan/aggregator sites.
# Tier 4 (3-4):  tabloids / heavy-aggregation / speculation.
# ---------------------------------------------------------------------------
CREDIBILITY = {
    "liverpoolfc.com": 10, "premierleague.com": 10, "uefa.com": 10,
    "bbc.co.uk": 10, "bbc.com": 10, "theathletic.com": 10, "nytimes.com": 10,
    "skysports.com": 9, "theguardian.com": 9, "reuters.com": 10, "apnews.com": 10,
    "espn.com": 8, "espn.co.uk": 8, "thetimes.co.uk": 9, "telegraph.co.uk": 8,
    "independent.co.uk": 7, "liverpoolecho.co.uk": 7, "goal.com": 6,
    "football.london": 6, "90min.com": 6, "fabrizioromano.com": 9,
    "thisisanfield.com": 6, "empireofthekop.com": 5, "anfieldwatch.co.uk": 5,
    "rousingthekop.com": 5, "caughtoffside.com": 4, "football365.com": 6,
    "teamtalk.com": 5, "mirror.co.uk": 4, "thesun.co.uk": 4, "dailymail.co.uk": 4,
    "tribuna.com": 5, "sportsmole.co.uk": 5, "givemesport.com": 5,
    "footballtransfers.com": 5, "metro.co.uk": 4,
}
DEFAULT_CREDIBILITY = 4  # unknown outlet

# Known reliable transfer/club reporters (many break news first on X/Twitter).
# When one is named in a story, we treat the report as high-accuracy regardless
# of which outlet republished it — these are the people the user trusts most.
TRUSTED_REPORTERS = [
    "fabrizio romano", "david ornstein", "ornstein", "james pearce",
    "paul joyce", "melissa reddy", "david lynch", "neil jones",
    "florian plettenberg", "ben jacobs", "simon hughes", "chris bascombe",
]

# ---------------------------------------------------------------------------
# 2. FEEDS
# Direct outlet feeds (fast, clean) + Google News search feeds (web-wide:
# pulls from hundreds of outlets and surfaces individual journalists' reports).
# Google News items carry their original source, so credibility is scored from
# the source domain via CREDIBILITY above.
# ---------------------------------------------------------------------------
DIRECT_FEEDS = [
    "https://feeds.bbci.co.uk/sport/football/teams/liverpool/rss.xml",
    "https://www.theguardian.com/football/liverpool/rss",
    "https://www.liverpoolfc.com/news.rss",
    "https://www.thisisanfield.com/feed/",
    "https://www.empireofthekop.com/feed/",
    "https://www.liverpoolecho.co.uk/all-about/liverpool-fc?service=rss",
]

# Google News RSS search queries — broaden coverage across the whole web.
GOOGLE_QUERIES = [
    "Liverpool FC transfer",
    "Liverpool FC Iraola",
    "Liverpool FC injury",
    "Liverpool FC Isak OR Wirtz OR Gakpo",
    '"Liverpool" Diomande OR Bouaddi OR transfer target',
    "Liverpool World Cup 2026 player",
    # Reporter-focused queries. Top reporters break news first on X/Twitter;
    # Google News surfaces their articles AND aggregators quoting their posts,
    # which is the practical way to capture social-media scoops without the
    # (paid, auth-walled) X API.
    'Liverpool "Fabrizio Romano"',
    'Liverpool "James Pearce"',
    'Liverpool "Paul Joyce"',
    'Liverpool "David Ornstein"',
    'Liverpool ("Melissa Reddy" OR "David Lynch" OR "Neil Jones")',
]
def google_news_url(q):
    return ("https://news.google.com/rss/search?q="
            + quote_plus(q + " when:21d")
            + "&hl=en-GB&gl=GB&ceid=GB:en")

# Relevance keywords (lowercase). Hits add to the relevance score.
KEYWORDS = [
    "liverpool", "anfield", "iraola", "transfer", "signing", "signs", "deal",
    "contract", "injury", "fixture", "wirtz", "isak", "salah", "van dijk",
    "szoboszlai", "ekitike", "gakpo", "mac allister", "gravenberch", "frimpong",
    "konate", "robertson", "diomande", "fsg", "premier league", "world cup",
]

MAX_ITEMS = 30
RECENCY_DAYS = 21


def clean(text):
    text = unescape(text or "")
    text = re.sub(r"<[^>]+>", "", text)
    return re.sub(r"\s+", " ", text).strip()


def published_dt(entry):
    for key in ("published_parsed", "updated_parsed"):
        t = entry.get(key)
        if t:
            return dt.datetime(*t[:6])
    return dt.datetime.utcnow()


def domain_of(entry, fallback=""):
    # Google News puts the outlet in entry.source.title; otherwise parse link.
    src = entry.get("source")
    if isinstance(src, dict) and src.get("href"):
        link = src["href"]
    else:
        link = entry.get("link", fallback)
    m = re.search(r"https?://([^/]+)/?", link or "")
    host = (m.group(1) if m else "").lower().replace("www.", "")
    return host


def credibility_for(host, text):
    score = DEFAULT_CREDIBILITY
    for dom, val in CREDIBILITY.items():
        if host.endswith(dom):
            score = val
            break
    # Trusted-reporter boost: when one of these reporters is named, treat the
    # story as high-accuracy (floor of 9) regardless of which outlet ran it,
    # since the information originates from a reliable, well-sourced journalist.
    if any(r in text for r in TRUSTED_REPORTERS):
        score = min(10, max(score + 1, 9))
    return score


def source_name(entry, host):
    src = entry.get("source")
    if isinstance(src, dict) and src.get("title"):
        return src["title"]
    pretty = {
        "bbc.co.uk": "BBC Sport", "theguardian.com": "The Guardian",
        "liverpoolfc.com": "Liverpool FC", "thisisanfield.com": "This Is Anfield",
        "empireofthekop.com": "Empire of the Kop", "liverpoolecho.co.uk": "Liverpool Echo",
        "skysports.com": "Sky Sports", "espn.com": "ESPN",
    }
    for dom, name in pretty.items():
        if host.endswith(dom):
            return name
    return host or "Unknown"


def relevance_score(title, summary, age_days):
    hay = (title + " " + summary).lower()
    rel = 0
    if "liverpool" in hay:
        rel += 40
    if "liverpool" in summary.lower():
        rel += 10
    hits = sum(1 for k in KEYWORDS if k in hay)
    rel += min(hits * 5, 30)
    if age_days <= 1:
        rel += 20
    elif age_days <= 3:
        rel += 12
    elif age_days <= 7:
        rel += 6
    return min(rel, 100)


def collect():
    feeds = list(DIRECT_FEEDS) + [google_news_url(q) for q in GOOGLE_QUERIES]
    now = dt.datetime.utcnow()
    items = []
    for url in feeds:
        try:
            feed = feedparser.parse(url)
        except Exception as e:
            print(f"  ! feed error {url[:50]}: {e}")
            continue
        label = "google-news" if "news.google" in url else url.split("//")[1][:28]
        n = 0
        for e in feed.entries:
            title = clean(e.get("title", ""))
            summary = clean(e.get("summary", ""))
            if not title:
                continue
            hay = (title + " " + summary).lower()
            if "liverpool" not in hay and "anfield" not in hay:
                continue
            pub = published_dt(e)
            age = (now - pub).days
            if age > RECENCY_DAYS or age < -1:
                continue
            host = domain_of(e, url)
            cred = credibility_for(host, hay)
            rel = relevance_score(title, summary, age)
            # Accuracy-weighted ranking: relevance scaled by source reliability.
            composite = round(rel * (0.45 + 0.55 * cred / 10), 1)
            items.append({
                "headline": title,
                "summary": summary[:240],
                "source": source_name(e, host),
                "credibility": cred,
                "relevance": rel,
                "rank_score": composite,
                "date": pub.strftime("%Y-%m-%d"),
                "url": e.get("link", ""),
            })
            n += 1
        print(f"  {label}: {n} items")
    return items


def dedupe_and_rank(items):
    items.sort(key=lambda x: x["rank_score"], reverse=True)
    seen, out = set(), []
    for it in items:
        key = re.sub(r"[^a-z0-9]", "", it["headline"].lower())[:55]
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    for i, it in enumerate(out, 1):
        it["rank"] = i
    return out[:MAX_ITEMS]


def load_reference():
    p = DATA / "reference.json"
    return json.loads(p.read_text(encoding="utf-8")) if p.exists() else {}


def write_excel(news, ref):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    red = PatternFill("solid", fgColor="C8102E")
    head = Font(bold=True, color="FFFFFF")

    def sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.fill = red; c.font = head; c.alignment = Alignment(vertical="center")
        for r in rows:
            ws.append(r)
        for i, h in enumerate(headers, 1):
            vals = [len(str(h))] + [len(str(r[i-1])) for r in rows] if rows else [len(str(h))]
            ws.column_dimensions[chr(64+i)].width = min(max(vals) + 3, 70)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    wb.remove(wb.active)
    sheet("News (ranked)",
          ["Rank", "Date", "Headline", "Source", "Credibility", "Relevance", "Rank score", "URL"],
          [[n["rank"], n["date"], n["headline"], n["source"], n["credibility"],
            n["relevance"], n["rank_score"], n["url"]] for n in news])
    sheet("Source credibility",
          ["Domain", "Credibility (1-10)"],
          [[d, v] for d, v in sorted(CREDIBILITY.items(), key=lambda x: -x[1])])
    if ref.get("squad"):
        sheet("Squad",
              ["#", "Player", "Position", "Slot", "Age", "Nat", "Contract", "Iraola fit", "Injury"],
              [[p.get("num"), p["name"], p["posGroup"], p.get("slot",""), p["age"], p["nat"],
                p["contract"], p.get("fit",""), (p.get("injury") or "")] for p in ref["squad"]])
    if ref.get("transfers"):
        t = ref["transfers"]; rows = []
        for c in t.get("in", []):    rows.append(["IN (confirmed)", c["who"], c["meta"], c["fee"], ""])
        for c in t.get("out", []):   rows.append(["OUT (confirmed)", c["who"], c["meta"], c["fee"], ""])
        for c in t.get("rumIn", []): rows.append(["Rumour IN", c["who"], c["meta"], c["fee"], c["p"]])
        for c in t.get("rumOut", []):rows.append(["Rumour OUT", c["who"], c["meta"], c["fee"], c["p"]])
        sheet("Transfers", ["Type", "Player", "Detail", "Fee", "Likelihood"], rows)
    if ref.get("injuries"):
        sheet("Injuries", ["Player", "Issue", "Status", "Severity"],
              [[i["player"], i["issue"], i["status"], i.get("severity","")] for i in ref["injuries"]])
    wb.save(DATA / "liverpool_data.xlsx")
    print(f"  wrote {DATA/'liverpool_data.xlsx'}")


# ---------------------------------------------------------------------------
# 3. LIVE TRANSFER RUMOURS — derived from the scraped feed every run, so the
# Transfers tab refreshes daily without manual editing. Direction (in/out) and
# likelihood are inferred from the wording and weighted by source credibility.
# ---------------------------------------------------------------------------
TRANSFER_KEYS = ["transfer", "sign", "signing", "signs", "signed", "deal", "bid",
    "fee", "release clause", "loan", "exit", "leav", "sold", "sale", "swoop",
    "agree", "agreed", "medical", "here we go", "linked", " link", "target",
    "wants", "interest", "talks", "suitor", "close to", "pursuit", "hijack", "move for"]
HIGH_WORDS = ["here we go", "confirmed", "confirm", "completes", "complete", "done deal",
    "agreed", "agree personal terms", "sealed", " seal ", "medical", "unveiled",
    "announce", "signs ", "signed", "triggered", "joins"]
MED_WORDS = ["close to", "advanced", "in talks", "verbal", "bid", "submitted",
    "negotiat", "accept", "agreement", "reach", "personal terms", "offer"]
OUT_WORDS = ["exit", "leav", "sold", "sale", "up for sale", "offer for", "bid for",
    "suitor", "wanted by", "to join", "departure", "sell ", "swap", "wants to leave"]


def squad_tokens(ref):
    toks = set()
    for grp in (ref.get("squad", []), ref.get("loans", [])):
        for p in grp:
            nm = (p.get("name") or "").replace(" (c)", "").strip()
            if nm:
                toks.add(nm.split()[-1].lower())
    for c in ref.get("transfers", {}).get("out", []):
        w = c.get("who", "")
        if w:
            toks.add(w.split()[-1].lower())
    return {x for x in toks if len(x) > 3}


def build_transfers(items, ref, limit=16):
    lfc = squad_tokens(ref)
    out = []
    for it in items:
        hay = (it["headline"] + " " + it.get("summary", "")).lower()
        if not any(k in hay for k in TRANSFER_KEYS):
            continue
        mentions_lfc = any(n in hay for n in lfc)
        d = "out" if (mentions_lfc and any(w in hay for w in OUT_WORDS)) else "in"
        if any(w in hay for w in HIGH_WORDS):
            p = "High"
        elif any(w in hay for w in MED_WORDS):
            p = "Medium"
        else:
            p = "Low"
        if it["credibility"] < 5 and p == "High":   # weak source can't assert "High"
            p = "Medium"
        out.append({"headline": it["headline"], "source": it["source"],
                    "credibility": it["credibility"], "date": it["date"],
                    "url": it["url"], "dir": d, "p": p, "rank_score": it["rank_score"]})
        if len(out) >= limit:
            break
    return out


def main():
    print("Scraping Liverpool news (direct feeds + Google News web-wide)...")
    news = dedupe_and_rank(collect())
    print(f"Ranked {len(news)} stories.")
    ref = load_reference()
    rumours = build_transfers(news, ref)
    print(f"Derived {len(rumours)} live transfer rumours.")
    payload = {
        "updated": dt.datetime.utcnow().strftime("%d %b %Y %H:%M UTC"),
        "count": len(news),
        "items": news,
        "transfers_live": rumours,
    }
    (DATA / "news.json").write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  wrote {DATA/'news.json'}")
    try:
        write_excel(news, ref)
    except Exception as e:
        print(f"  ! Excel write failed: {e}")
    print("Done.")


if __name__ == "__main__":
    main()
